import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def browse_url(url):
    with urllib.request.urlopen(url) as f:
        return f.read().decode('utf-8')
    
def format_url(url):
    if (url.startswith('https')):
        return url
    return 'https://dev.java' + url

def apply_cell_color(cell, color="FFFFFF"):
    cell.fill = PatternFill(patternType='solid', fgColor=color)

def apply_cell_border(cell, color="000000"):
    line = Side(border_style="thin", color=color)
    full_border = Border(
        left=line,
        right=line,
        top=line,
        bottom=line
    )
    cell.border = full_border

def apply_cell_text_center(cell):
   cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_cell_text_style(cell, color="000000", bold=False, size=12):
    cell.font = Font(name="Arial", color=color, bold=bold, size=size)

def format_table_header_cell(cell):
    apply_cell_border(cell)
    apply_cell_text_center(cell)
    apply_cell_text_style(cell, size=12, bold=True)

def format_main_category(cell):
    apply_cell_color(cell, color="9fc5e8")
    apply_cell_border(cell)
    apply_cell_text_center(cell)
    apply_cell_text_style(cell, size=12)

def format_sub_category(cell):
    apply_cell_color(cell, color="ffe599")
    apply_cell_border(cell)
    apply_cell_text_center(cell)
    apply_cell_text_style(cell, size=10)

def format_sub_sub_category(cell):
    apply_cell_color(cell, color="dde599")
    apply_cell_border(cell)
    apply_cell_text_center(cell)
    apply_cell_text_style(cell, size=8)

def format_cell_link(cell, link):
    cell.value = link
    cell.hyperlink = link

def main():
    wb = Workbook()
    
    # Pre-formatting
    ws = wb.active
    ws.title = "Java tutorials"
    # Header formatting
    header = ws['A1']
    ws['A1'] = "Dev.Java Courses Tracking"
    
    apply_cell_border(header)
    apply_cell_color(header, color="000000")
    apply_cell_text_center(header)
    apply_cell_text_style(header, color="FFFFFF", size=24, bold=True)

    ws.merge_cells('A1:Z3')

    # Header space
    ws.merge_cells('A4:Z5')

    # Table header
    ws['A6'] = 'Tutorial'
    format_table_header_cell(ws['A6'])
    ws['B6'] = 'Link'
    format_table_header_cell(ws['B6'])
    ws['C6'] = 'Progress'
    format_table_header_cell(ws['C6'])

    # Meta-data
    dv = DataValidation(
        type="list",
        formula1='"NOT STARTED","IN PROGRESS","DONE"',
        allow_blank=False,
    )
    ws.add_data_validation(dv)

    # Fill with contents
    cell_pointer = 7
    format_sub_category(ws['A7'])
    ws['A7'] = "Test"
    ws.merge_cells('A7:C8')

    # Fetching content
    html_contents = browse_url('https://dev.java/learn/')
    soup = BeautifulSoup(html_contents, 'lxml')
    result = {}

    # Main ctagories
    categories_div = soup.find_all('div', 'learn-group')
    for div in categories_div:
        title = div.h2
        result[title.string] = []
        
        # Categories links
        categories_links = div.find_all('a')
        for cl in categories_links:
            new_lesson = { "link": format_url(cl.get('href')), "text": cl.string, "children": []}
            new_lesson_content = BeautifulSoup(browse_url(new_lesson["link"]), "lxml")
            tg = new_lesson_content.find(id='tutorial-group')
            if (tg):
                tgl = tg.find_all("a")
                for tgla in tgl:
                    new_lesson["children"].append({ "link": format_url(tgla.get('href')), "text": tgla.string })

            result[title.string].append(new_lesson)

    # Formatting based on the fetched contents
    for k, v in result.items():
        # Formatting main category
        current_cell_index = 'A' + str(cell_pointer)
        ws[current_cell_index] = k
        category_cell = ws[current_cell_index]
        format_main_category(category_cell)
        ws.merge_cells(current_cell_index + ':' + 'C' + (str(cell_pointer + 1)))
        cell_pointer += 2

        for link_index, link in enumerate(v):
            is_sub = len(link["children"]) > 0
            previous = None
            if link_index > 0:
                previous = v[link_index - 1]

            if is_sub:
                # Formatting sub header 
                current_cell_index = 'A' + str(cell_pointer)
                ws[current_cell_index] = link["text"]
                sub_category_cell = ws[current_cell_index]
                format_sub_category(sub_category_cell)
                ws.merge_cells(current_cell_index + ':' + 'C' + (str(cell_pointer + 1)))
                cell_pointer += 2

                # Links
                for c in link["children"]:
                    ws['A' + str(cell_pointer)] = c["text"]
                    format_cell_link(ws['B' + str(cell_pointer)], c["link"])
                    dv.add(ws['C' + str(cell_pointer)])
                    ws['C' + str(cell_pointer)] = "NOT STARTED"
                    cell_pointer +=1
            else:
                if previous:
                    previous_is_sub = len(previous["children"]) > 0
                    if previous_is_sub:
                        # Add blank space
                        current_cell_index = 'A' + str(cell_pointer)
                        ws.merge_cells(current_cell_index + ':' + 'C' + (str(cell_pointer + 1)))
                        cell_pointer += 2

                ws['A' + str(cell_pointer)] = link["text"]
                format_cell_link(ws['B' + str(cell_pointer)], link["link"])
                dv.add(ws['C' + str(cell_pointer)])
                ws['C' + str(cell_pointer)] = "NOT STARTED"
                cell_pointer +=1
    # Write file
    wb.save("java-learning-tracking.xlsx")


if __name__ == "__main__":
    main()
