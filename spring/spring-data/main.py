from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment # Import nécessaire pour l'indentation
from reference import format_table_header_cell, format_main_category, format_sub_category, format_cell_link, format_sub_sub_category
import urllib.request

# --- Constante pour définir le niveau maximum qui peut être une catégorie ---
MAX_CATEGORY_LEVEL = 2 
# -------------------------------------------------------------------------

def fetch_html(url):
    with urllib.request.urlopen(url) as f:
        return f.read().decode('utf-8')

def format_url(href, base_url):
    if href.startswith('http'):
        return href
    if href.startswith('/'):
        return base_url.rstrip('/') + href
    return base_url.rstrip('/') + '/' + href

def extraire_structure_nav(soup, base_url):
    """
    Extrait la structure de navigation et indique si chaque élément a des enfants.
    Retourne: liste de tuples (niveau, libelle, lien, a_des_enfants)
    """
    nav = soup.find('nav', class_='nav-menu')
    if not nav:
        return []
    result = []
    
    def parcourir(ul, niveau=0):
        for li in ul.find_all('li', class_='nav-item', recursive=False):
            a = li.find('a', class_='nav-link')
            libelle = a.get_text(strip=True) if a else li.get_text(strip=True)
            lien = format_url(a['href'], base_url) if a and a.has_attr('href') else ''
            
            sous_ul = li.find('ul', class_='nav-list')
            a_des_enfants = sous_ul is not None
            
            result.append((niveau, libelle, lien, a_des_enfants))
            
            if sous_ul:
                parcourir(sous_ul, niveau + 1)
                
    ul_principale = nav.find('ul', class_='nav-list')
    if ul_principale:
        parcourir(ul_principale)
        
    return result

def inserer_separateur(ws, current_row, hauteur=15):
    """
    Insère une ligne vide pour servir de séparateur visuel.
    """
    ws.row_dimensions[current_row].height = hauteur
    return current_row + 1

def generer_excel_structure(structure, excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Suivi Spring Data'
    
    # En-têtes et formats de base
    ws['A1'] = "Suivi de progression - Spring Data"
    from reference import apply_cell_border, apply_cell_color, apply_cell_text_center, apply_cell_text_style
    apply_cell_border(ws['A1'])
    apply_cell_color(ws['A1'], color="000000")
    apply_cell_text_center(ws['A1'])
    apply_cell_text_style(ws['A1'], color="FFFFFF", size=18, bold=True)
    ws.merge_cells('A1:C2')
    ws.merge_cells('A3:C3')
    
    ws['A4'] = 'Libellé'
    format_table_header_cell(ws['A4'])
    ws['B4'] = 'Lien'
    format_table_header_cell(ws['B4'])
    ws['C4'] = 'Statut'
    format_table_header_cell(ws['C4'])
    
    dv = DataValidation(type="list", formula1='"À faire,En cours,Fait"', allow_blank=False)
    ws.add_data_validation(dv)
    
    # Remplissage des données avec gestion dynamique des niveaux
    row = 5
    
    # Suivi du niveau du dernier élément traité pour savoir s'il faut ajouter un séparateur
    dernier_niveau_traite = -1 
    
    for niveau, libelle, lien, a_des_enfants in structure:
        
        # LOGIQUE DE SÉPARATION :
        # On insère une ligne vide AVANT un nouvel élément de haut niveau (Niveau 0 ou 1)
        # s'il ne s'agit pas du tout premier élément du tableau.
        if (niveau <= 1) and (row > 5):
            # Le séparateur est ajouté AVANT le nouvel élément de haut niveau
            row = inserer_separateur(ws, row) 

        ws[f'A{row}'] = libelle
        
        # Condition de Regroupement : Est-ce un niveau de catégorie ET a-t-il des enfants ?
        if niveau <= MAX_CATEGORY_LEVEL and a_des_enfants:
            # --- Cas 1 : Niveau de Regroupement (Catégorie avec Enfants) ---
            
            # Application des formats spécifiques (Niveau 0, 1, 2, ...)
            if niveau == 0:
                format_main_category(ws[f'A{row}'])
            elif niveau == 1:
                format_sub_category(ws[f'A{row}'])
            else: # niveaux 2 et supérieurs qui sont des regroupements
                format_sub_sub_category(ws[f'A{row}']) 
                
            ws[f'A{row}'].hyperlink = lien
            ws.merge_cells(f'A{row}:C{row}')
            
        else:
            # --- Cas 2 : Élément de Contenu (Détail à Suivre) ---
            
            # Application d'une indentation visuelle pour la lisibilité
            # On utilise max(0, niveau-1) pour réduire légèrement l'indentation
            # si niveau 2 n'est pas fusionné, il a l'indentation 1
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=niveau) 
            
            format_cell_link(ws[f'B{row}'], lien)
            dv.add(ws[f'C{row}'])
            ws[f'C{row}'] = "À faire"
            
        dernier_niveau_traite = niveau # Mise à jour du dernier niveau traité
        row += 1
        
    # Optionnel: Ajouter un séparateur à la toute fin pour terminer le tableau proprement
    row = inserer_separateur(ws, row, hauteur=10)
    
    wb.save(excel_file)
    print(f"Fichier Excel généré : {excel_file}")

if __name__ == '__main__':
    # Étape 2 : lecture distante
    html = fetch_html('https://docs.spring.io/spring-data/commons/reference/')
    soup = BeautifulSoup(html, 'html.parser')
    structure = extraire_structure_nav(soup, 'https://docs.spring.io/spring-data/commons/reference/')
    generer_excel_structure(structure, 'spring-data-tracking.xlsx')
